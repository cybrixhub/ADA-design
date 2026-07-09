"""
Extract project data from ./scratch_extract/Website/ into:
  - public/projects/<slug>/*.jpg  (optimized)
  - lib/projects.ts               (typed catalog)
"""
from __future__ import annotations
import os
import re
import json
import shutil
from pathlib import Path
from docx import Document
from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "scratch_extract" / "Website"
OUT_IMAGES = ROOT / "public" / "projects"
OUT_TS = ROOT / "lib" / "projects.ts"

CATALOG = ROOT / "scratch_extract" / "Website" / "Projects for website.xlsx"

MAX_WIDTH = 1600
JPEG_QUALITY = 82


def slugify(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def parse_docx(path: Path) -> dict:
    """Extract TITLE and DESCRIPTION and any option / room lists."""
    doc = Document(path)
    lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    joined = "\n".join(lines)

    title = ""
    description = ""

    # TITLE: XYZ  or  TITLE:\nXYZ
    m = re.search(r"TITLE\s*:?\s*(.+?)(?:\n|$)", joined, re.IGNORECASE)
    if m:
        cand = m.group(1).strip()
        if cand.upper() == cand and len(cand) < 4:  # avoid picking up label word only
            m2 = re.search(r"TITLE\s*:?\s*\n\s*(.+?)(?:\n|$)", joined, re.IGNORECASE)
            if m2:
                cand = m2.group(1).strip()
        title = cand

    # DESCRIPTION: block up to next ALL-CAPS heading or double newline
    m = re.search(
        r"DESCRIPTION\s*:?\s*(.+?)(?:\n\s*[A-Z][A-Z\s\d/&\-]{5,}\s*(?:\n|$)|\Z)",
        joined,
        re.IGNORECASE | re.DOTALL,
    )
    if m:
        description = re.sub(r"\s+", " ", m.group(1)).strip()

    return {"title": title, "description": description, "raw": joined}


def load_categories() -> dict[str, str]:
    """Return {address: category}. Address matches the folder name (or close)."""
    wb = load_workbook(CATALOG, data_only=True)
    ws = wb["Sheet1"]
    current_category = None
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        if not cells:
            continue
        first_val = row[0]
        second_val = row[1] if len(row) > 1 else None
        # Category header: second col has text, first col is empty
        if first_val is None and second_val and isinstance(second_val, str):
            if second_val.strip() not in ("Project Grouping", "Living", "Bedroom", "Dining/Kitchen"):
                current_category = second_val.strip()
        elif isinstance(first_val, (int, float)) and second_val and current_category:
            mapping[second_val.strip()] = current_category
    return mapping


def optimize_image(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(src) as im:
        if im.mode in ("RGBA", "P"):
            bg = Image.new("RGB", im.size, (255, 255, 255))
            if im.mode == "RGBA":
                bg.paste(im, mask=im.split()[3])
            else:
                bg.paste(im.convert("RGBA"))
            im = bg
        elif im.mode != "RGB":
            im = im.convert("RGB")
        w, h = im.size
        if w > MAX_WIDTH:
            new_h = int(h * MAX_WIDTH / w)
            im = im.resize((MAX_WIDTH, new_h), Image.LANCZOS)
        im.save(dst, "JPEG", quality=JPEG_QUALITY, optimize=True, progressive=True)


def find_project_images(folder: Path) -> list[Path]:
    exts = {".jpg", ".jpeg", ".png"}
    imgs = []
    for root, _, files in os.walk(folder):
        for f in files:
            p = Path(root) / f
            if p.suffix.lower() in exts:
                imgs.append(p)
    imgs.sort()
    return imgs


def ts_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def main() -> None:
    if OUT_IMAGES.exists():
        shutil.rmtree(OUT_IMAGES)
    OUT_IMAGES.mkdir(parents=True)
    (ROOT / "lib").mkdir(exist_ok=True)

    cat_map = load_categories()

    projects = []
    for folder in sorted(SRC.iterdir()):
        if not folder.is_dir():
            continue
        address = folder.name.strip()
        slug = slugify(address)
        category = cat_map.get(address, "Uncategorized")

        # Find the docx
        docxes = list(folder.rglob("*.docx"))
        info = {"title": address, "description": ""}
        if docxes:
            parsed = parse_docx(docxes[0])
            if parsed["title"]:
                info["title"] = parsed["title"]
            info["description"] = parsed["description"]

        # Optimize images
        src_imgs = find_project_images(folder)
        images: list[str] = []
        for i, src in enumerate(src_imgs):
            dst_name = f"{i + 1:02d}.jpg"
            dst = OUT_IMAGES / slug / dst_name
            try:
                optimize_image(src, dst)
                images.append(f"/projects/{slug}/{dst_name}")
            except Exception as e:
                print(f"  ! skipped {src.name}: {e}")

        projects.append({
            "slug": slug,
            "title": info["title"],
            "address": address,
            "category": category,
            "description": info["description"],
            "images": images,
        })
        print(f"✓ {address} → {len(images)} imgs, cat={category}")

    # Write TS
    OUT_TS.parent.mkdir(exist_ok=True)
    lines = [
        "// AUTO-GENERATED by scripts/build_projects.py — do not edit by hand.",
        "export type Project = {",
        "  slug: string;",
        "  title: string;",
        "  address: string;",
        "  category: string;",
        "  description: string;",
        "  images: string[];",
        "};",
        "",
        "export const projects: Project[] = " + json.dumps(projects, indent=2, ensure_ascii=False) + ";",
        "",
        "export const categories: string[] = Array.from(new Set(projects.map(p => p.category)));",
        "",
    ]
    OUT_TS.write_text("\n".join(lines), encoding="utf-8")

    total_imgs = sum(len(p["images"]) for p in projects)
    print(f"\nDone. {len(projects)} projects, {total_imgs} images. TS → {OUT_TS.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
